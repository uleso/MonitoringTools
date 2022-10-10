#seg Pre,Post Node dağıtıcı
from openpyxl import Workbook, load_workbook

segLayout = ["D","G","J","M",
 "S","V","Y","AB","AE","AH","AK","AN",
 "AU","AX","BA","BD","BH","BL","BP","BT","BX","CB","CF","CJ"]
segPreLayout = ["D","G","J","M","AU","AX","BA","BD"]
segPostLayout = ["S","V","Y","AB","AE","AH","AK","AN","BH","BL","BP","BT","BX","CB","CF","CJ"]  #postta 8 node için ayarlı
wb = Workbook() 
wb = load_workbook(filename = 'Test NCT.xlsx')
ws = wb.active

sheets = wb.worksheets
def calcPreNodes():
    for sheet in sheets:
        for precol in sheet.iter_cols(2,2,2,100,True):
            return precol 
def calcPostNodes():       
    for sheet in sheets:
        for postcol in sheet.iter_cols(3,3,2,100,True):
            return postcol
def coordinateFinder(Layout,rowMarker,mode,nodePair):
    coordinateList = []
    if(mode == "Pre"):
        if(nodePair == 1 ):
            for colname in Layout[:1] + Layout[4:5]:
                coordinateList.append( colname + str(rowMarker)) 
        
        elif(nodePair == 2 ):
            for colname in Layout[:2]+ Layout[4:6]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 3 ):
            for colname in Layout[:3]+ Layout[4:7]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 4 ):
            for colname in Layout[:4]+ Layout[4:8]:
                coordinateList.append( colname + str(rowMarker)) 
        return coordinateList            
    elif(mode == "Post"):
        if(nodePair == 1 ):
            for colname in Layout[:1] + Layout[8:9]:
                coordinateList.append( colname + str(rowMarker)) 
        
        elif(nodePair == 2 ):
            for colname in Layout[:2]+ Layout[8:10]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 3 ):
            for colname in Layout[:3]+ Layout[8:11]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 4 ):
            for colname in Layout[:4]+ Layout[8:12]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 5 ):
            for colname in Layout[:5]+ Layout[8:13]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 6 ):
            for colname in Layout[:6]+ Layout[8:14]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 7 ):
            for colname in Layout[:7]+ Layout[8:15]:
                coordinateList.append( colname + str(rowMarker))
        elif(nodePair == 8 ):
            for colname in Layout[:8]+ Layout[8:16]:
                coordinateList.append( colname + str(rowMarker))
    return coordinateList
    
def parsePreNodes(nodeList):
    rowMarker = 2
    
    for cell in nodeList:
        mode = "Pre"
        #OLD NEW NODE İÇİN BİR KOŞUL DAHA YAZ
        if(type(cell) is int):
            cAddress = coordinateFinder(segPreLayout,rowMarker,mode,1)
            for address in cAddress:
                ws[address]= cell                        
        elif(cell.count("/") == 1):
            pass
        elif(cell.count(',') == 1):
            node1=cell[:6]
            node2=cell[8:14]
            cAddress = coordinateFinder(segPreLayout,rowMarker,mode,2)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[2])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[3])
            for address in node1Address:
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2
        elif(cell.count(',')== 2):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            cAddress = coordinateFinder(segPreLayout,rowMarker,mode,3)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[3])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[4])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[5])
            
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2
            for address in node3Address:
                ws[address] = node3
        elif(cell.count(',')== 3):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            cAddress = coordinateFinder(segPreLayout,rowMarker,mode,4)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[4])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[5])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[6])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[7])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
        rowMarker += 1

def parsePostNodes(nodeList):
    mode ="Post"
    rowMarker = 2
    
    for cell in nodeList:
        if(type(cell) is int):
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,1)
            for address in cAddress:
                ws[address]= cell                        
        elif(cell.count("/") == 1):
            pass        
        elif(cell.count(',') == 1):
            node1=cell[:6]
            node2=cell[8:14]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,2)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[2])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[3])
            for address in node1Address:
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2
        elif(cell.count(',')== 2):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,3)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[3])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[4])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[5])
            
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2
            for address in node3Address:
                ws[address] = node3
        elif(cell.count(',')== 3):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,4)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[4])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[5])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[6])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[7])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
        elif(cell.count(',')== 4):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            node5=cell[32:38]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,5)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[5])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[6])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[7])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[8])
            node5Address = []
            node5Address.append(cAddress[4])
            node5Address.append(cAddress[9])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
            for address in node5Address:
                ws[address] = node5
        elif(cell.count(',')== 5):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            node5=cell[32:38]
            node6=cell[40:46]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,6)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[6])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[7])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[8])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[9])
            node5Address = []
            node4Address.append(cAddress[4])
            node4Address.append(cAddress[10])
            node6Address = []
            node6Address.append(cAddress[5])
            node6Address.append(cAddress[11])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
            for address in node5Address:
                ws[address] = node5
            for address in node6Address:
                ws[address] = node6
        elif(cell.count(',')== 6):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            node5=cell[32:38]
            node6=cell[40:46]
            node7=cell[48:54]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,7)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[7])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[8])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[9])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[10])
            node5Address = []
            node5Address.append(cAddress[4])
            node5Address.append(cAddress[11])
            node6Address = []
            node6Address.append(cAddress[5])
            node6Address.append(cAddress[12])
            node7Address = []
            node7Address.append(cAddress[5])
            node7Address.append(cAddress[13])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
            for address in node5Address:
                ws[address] = node5
            for address in node6Address:
                ws[address] = node6
            for address in node7Address:
                ws[address] = node7
        elif(cell.count(',')== 7):
            node1=cell[:6]
            node2=cell[8:14]
            node3=cell[16:22]
            node4=cell[24:30]
            node5=cell[32:38]
            node6=cell[40:46]
            node7=cell[48:54]
            node8=cell[56:62]
            cAddress = coordinateFinder(segPostLayout,rowMarker,mode,8)
            node1Address = []
            node1Address.append(cAddress[0])
            node1Address.append(cAddress[8])
            node2Address = []
            node2Address.append(cAddress[1])
            node2Address.append(cAddress[9])
            node3Address = []
            node3Address.append(cAddress[2])
            node3Address.append(cAddress[10])
            node4Address = []
            node4Address.append(cAddress[3])
            node4Address.append(cAddress[11])
            node5Address = []
            node5Address.append(cAddress[4])
            node5Address.append(cAddress[12])
            node6Address = []
            node6Address.append(cAddress[5])
            node6Address.append(cAddress[13])
            node7Address = []
            node7Address.append(cAddress[5])
            node7Address.append(cAddress[14])
            node8Address = []
            node8Address.append(cAddress[5])
            node8Address.append(cAddress[15])
            for address in node1Address:        
                ws[address] = node1
            for address in node2Address:
                ws[address] = node2    
            for address in node3Address:
                ws[address] = node3
            for address in node4Address:
                ws[address] = node4
            for address in node5Address:
                ws[address] = node5
            for address in node6Address:
                ws[address] = node6
            for address in node7Address:
                ws[address] = node7
            for address in node8Address:
                ws[address] = node8
        rowMarker += 1
filterPreNode = calcPreNodes()
filteredPreNode = tuple(filter(lambda item: item != None,filterPreNode))
parsePreNodes(filteredPreNode)
filterPostNode = calcPostNodes()
filteredPostNode = tuple(filter(lambda item:item != None,filterPostNode))
parsePostNodes(filteredPostNode)
wb.save("YargıDağıtıldı.xlsx")
